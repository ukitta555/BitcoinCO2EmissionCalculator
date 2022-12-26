import logging
from datetime import datetime
from pprint import pprint

import django
import openpyxl
from django.db import transaction
from openpyxl import load_workbook

from src.bitcoin_emissions.consts import CLOUDFLARE_LOCATION_DATA, CLOUDFLARE_REGION, CURRENT_INFO_SHEET_NAME

# from src.bitcoin_emissions.models import Pool, Location, PoolLocation

logger = logging.getLogger(__name__)

def print_pools():

    wb = load_workbook(filename='../data/Pool_data_final.xlsx')
    def get_pool_names_column(sheet):
        return sheet['A']

    unique_pools = set()
    for sheet in wb:
        for cell in get_pool_names_column(sheet):
            if cell.value:
                unique_pools.add(cell.value)
    pprint(list(sorted(unique_pools)))


def parse_excel_for_pool_and_location_info(workbook: openpyxl.Workbook):
    from src.bitcoin_emissions.models import Pool, Location, PoolLocation

    def server_is_hosted_by_cloudflare(location_name):
        return location_name == CLOUDFLARE_REGION

    with transaction.atomic():
        for sheet in workbook:
            # logger.info(f"Working with sheet {sheet.title}")
            print(f"Working with sheet {sheet.title}")
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                pool_name, location, emission_factor, source, \
                    latitude, longitude = row
                if sheet.title == CURRENT_INFO_SHEET_NAME:
                    validity_date = datetime.today()
                else:
                    validity_date = datetime.strptime(sheet.title, "%b %d %Y")

                if server_is_hosted_by_cloudflare(location):
                    location, latitude, longitude = CLOUDFLARE_LOCATION_DATA

                pool, new_pool_was_created = Pool.objects.get_or_create(pool_name=pool_name)
                location, new_location_was_created = Location.objects.get_or_create(
                    location_name=location,
                    latitude=latitude,
                    longitude=longitude,
                )
                pool_location_for_a_specific_date = PoolLocation.objects.create(
                    blockchain_pool=pool,
                    blockchain_pool_location=location,
                    valid_for_date=validity_date,
                    emission_factor=emission_factor,
                    information_source=source
                )
                # logger.info(f"Finished row {row_index} of sheet {sheet.title}")
                print(f"Finished row {row_index} of sheet {sheet.title}")

if __name__ == '__main__':
    django.setup()
    wb = load_workbook(filename='../data/Pool_data_final.xlsx')
    parse_excel_for_pool_and_location_info(wb)