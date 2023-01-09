import logging
from datetime import datetime

from django.http import HttpResponseBadRequest
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView

from src.bitcoin_emissions.calculations.metrics_calculation_runner import MetricsCalculationRunner
from src.bitcoin_emissions.management.commands.parse_excel_data import Command
from src.bitcoin_emissions.models import PoolElectricityConsumptionAndCO2EEmissionHistory, Pool, Location, PoolLocation, \
    BitcoinDifficulty, MiningGear, BlocksFoundByPoolPerWindow
from src.bitcoin_emissions.models.uuid_base_db_model import UUIDModel
from src.bitcoin_emissions.serializers import EmissionSerializer
from src.bitcoin_emissions.xlsx_data_parser import ExcelParser

logger = logging.getLogger(__name__)

class Co2AndElectricityView(APIView):

    def get(self, request):
        try:
            start_date = datetime.strptime(request.GET.get('start', '2021-01-01'), '%Y-%m-%d')
            end_date = datetime.strptime(request.GET.get('end', '2021-01-01'), '%Y-%m-%d')
            if not datetime(year=2021, month=1, day=1) <= start_date <= end_date <= datetime.today():
                raise Exception
        except Exception as e:
            return HttpResponseBadRequest(content="Bad request, please check whether you provided dates in a correct "
                                                  "format (YYYY-MM-DD)")
        try:
            result = \
                PoolElectricityConsumptionAndCO2EEmissionHistory\
                .objects\
                .filter(
                    date__gte=start_date,
                    date__lte=end_date
                ).select_related(
                    'location_of_servers'
                )
            serializer = EmissionSerializer(result, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.exception(e)
            return HttpResponseBadRequest(content="Request failed while fetching the data.")


class StartFetchingView(APIView):

    def get(self, request):
        MetricsCalculationRunner.calculate_metrics_up_until_today(
            start_date=datetime(year=2021, month=1, day=1)
        )
        return Response(status=200)


class ImportPoolData(APIView):

    def get(self, request):
        Command().handle()
        # pool_data = load_workbook(filename="./data/Pool_data_final.xlsx")
        # ExcelParser.parse_excel_for_pool_and_location_info(workbook=pool_data)
        # mining_gear_data = load_workbook(filename='./data/Antminer Models[90].xlsx')
        # ExcelParser.parse_excel_with_mining_gear_data(workbook=mining_gear_data)
        return Response(status=200)


class CleanAllTables(APIView):

    def get(self, request):
        PoolElectricityConsumptionAndCO2EEmissionHistory.objects.all().delete()
        logger.info("Removed all history")
        BlocksFoundByPoolPerWindow.objects.all().delete()
        logger.info("Removed all window info")
        PoolLocation.objects.all().delete()
        logger.info("Removed pool location window info")
        Location.objects.all().delete()
        logger.info("Removed location  info")
        Pool.objects.all().delete()
        logger.info("Removed pool info")
        BitcoinDifficulty.objects.all().delete()
        logger.info("Removed bitcoin difficulty info")
        MiningGear.objects.all().delete()
        logger.info("Removed mining gear info")
        UUIDModel.objects.all().delete()
        logger.info("Removed UUID info")


