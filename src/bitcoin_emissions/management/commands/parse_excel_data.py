from django.core.management import BaseCommand
from openpyxl import load_workbook

from src.bitcoin_emissions.xlsx_data_parser import ExcelParser


class Command(BaseCommand):
    def handle(self, **options):
        pool_data = load_workbook(filename="./data/Pool_data_final.xlsx")
        ExcelParser.parse_excel_for_pool_and_location_info(workbook=pool_data)
        mining_gear_data = load_workbook(filename='./data/Antminer Models[90].xlsx')
        ExcelParser.parse_excel_with_mining_gear_data(workbook=mining_gear_data)
