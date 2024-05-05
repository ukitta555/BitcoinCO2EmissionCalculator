import csv
from datetime import datetime
from io import BytesIO
from tempfile import NamedTemporaryFile
from venv import logger
from django import forms
from django.contrib import admin
from django.db.models.functions import Lower
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from openpyxl import Workbook, load_workbook
import pandas as pd

from src.bitcoin_emissions.calculations.metrics_calculation_runner import MetricsCalculationRunner
from src.bitcoin_emissions.models.average_efficiency_db_model import AverageEfficiency
from src.bitcoin_emissions.models.bitcoin_difficulty_db_model import BitcoinDifficulty
from src.bitcoin_emissions.models.blocks_found_by_pool_per_window_db_model import BlocksFoundByPoolPerWindow
from src.bitcoin_emissions.models.co2_electricity_history_per_server_db_model import CO2ElectricityHistoryPerServer
from src.bitcoin_emissions.models.hashrate_per_pool_server_db_model import HashRatePerPoolServer
from src.bitcoin_emissions.models.location_db_model import Location
from src.bitcoin_emissions.models.mining_gear_db_model import MiningGear
from src.bitcoin_emissions.models.network_hashrate_db_model import NetworkHashRate
from src.bitcoin_emissions.models.pool_db_model import Pool
from src.bitcoin_emissions.models.pool_electricity_consumption_and_co2e_emission_history_db_model import PoolElectricityConsumptionAndCO2EEmissionHistory
from src.bitcoin_emissions.models.pool_locations_db_model import PoolLocation
from src.bitcoin_emissions.models.uuid_base_db_model import UUIDModel
from src.bitcoin_emissions.serializers import EmissionSerializer
from src.bitcoin_emissions.xlsx_data_parser import ExcelParser


class ExcelImportForm(forms.Form):
    excel_file = forms.FileField()

class DateRangeForm(forms.Form):
    start_date = forms.DateField()
    end_date = forms.DateField()


# Register your models here.
@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    def get_ordering(self, request):
        return [Lower('location_name')]

@admin.register(Pool)
class PoolAdmin(admin.ModelAdmin):
    def get_ordering(self, request):
        return [Lower('pool_name')]

@admin.register(MiningGear)
class MiningGearAdmin(admin.ModelAdmin):
    change_list_template = "entities/change_list_import_excel.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
        ]
        return my_urls + urls

    def import_excel(self, request):
        if request.user.is_authenticated:
            if request.method == "POST":
                excel_file = request.FILES["excel_file"]
                pool_data = load_workbook(excel_file.file)
                # Create database objects from passed in data
                ExcelParser.parse_excel_with_mining_gear_data(workbook=pool_data)
                self.message_user(request, "Your Excel file has been imported")
                return redirect("..")
            form = ExcelImportForm()
            payload = {"form": form}
            return render(
                request, "admin/excel_form.html", payload
            )
        else:
            return redirect("../..")

    def get_ordering(self, request):
        return ['-release_date']  


@admin.register(PoolElectricityConsumptionAndCO2EEmissionHistory)
class LocationEmissionHistoryAdmin(admin.ModelAdmin):
    date_hierarchy = 'date'
    change_list_template = "entities/change_list_emission_history.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('clear_all_data/', self.clear_all_data),
            path('calculate_emissions/', self.calculate_emissions),
            path('export_excel_per_location/', self.export_data)
        ]
        return my_urls + urls

    def export_data(self, request):
        if request.user.is_authenticated:
            meta = PoolElectricityConsumptionAndCO2EEmissionHistory._meta
            columns = [
                'date', 
                'total_electricity_usage_for_location_at_this_date',
                'total_co2e_emissions_for_location_at_this_date',
                'location_of_servers.latitude',
                'location_of_servers.longitude',  
                'location_of_servers.location_name', 
                'is_cloudflare',
                'averaged_difficulty',
                'network_hash_rate_720_block_window',
                'averaged_gear_efficiency',
                'ServerData.blockchain_pool_name',
                'ServerData.hash_rate_at_this_date',
                'ServerData.electricity_usage_at_this_date',
                'ServerData.co2e_emissions_at_this_date',
            ]

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(meta)
            
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Emissions data"
            sheet.append(columns)
            serialized_data = EmissionSerializer(PoolElectricityConsumptionAndCO2EEmissionHistory.objects.all().select_related('location_of_servers'), many=True)        
            normalized_df = pd.json_normalize(
                serialized_data.data, 
                'servers_at_location', 
                [
                    'date',
                    'electricity_usage',
                    'co2e_emissions',
                    [
                        'location_of_servers',
                        'latitude'
                    ],
                    [
                        'location_of_servers',
                        'longitude'
                    ],
                    [
                        'location_of_servers',
                        'location_name'
                    ],
                    'is_cloudflare',
                    'averaged_difficulty',
                    'network_hash_rate_720_block_window',
                    'averaged_gear_efficiency'
                ],
                record_prefix="ServerData."
            )
            
            columns_list = normalized_df.columns.to_list()
            columns_list = columns_list[4:] + columns_list[0:4]

            normalized_df = normalized_df[columns_list]

            for obj in normalized_df.iterrows():
                row = list(obj[1])
                sheet.append(row)


            with NamedTemporaryFile() as tmp:
                wb.save(tmp.name)
                response.write(tmp.read())


            self.message_user(request, "All emission records (grouped by location) have been exported to an Excel file. Please check your downloads folder!")
            return response
        else:
            return redirect("../..")
        
    def calculate_emissions(self, request):
        if request.user.is_authenticated:
            if request.method == "POST":
                start_date = datetime.strptime(request.POST.get('start_date', '2021-01-01'), '%Y-%m-%d')
                end_date = datetime.strptime(request.POST.get('end_date', '2021-01-01'), '%Y-%m-%d')
            
                logger.info(f"Start date {start_date}: ")
                PoolElectricityConsumptionAndCO2EEmissionHistory.objects.filter(date=start_date).delete()
                logger.info("Removed location history info")
                CO2ElectricityHistoryPerServer.objects.filter(date=start_date).delete()
                logger.info("Removed pool history info")
                HashRatePerPoolServer.objects.filter(date=start_date).delete()
                logger.info("Removed hash rate per server info")
                AverageEfficiency.objects.filter(date=start_date).delete()
                logger.info("Removed average efficiency info")
                NetworkHashRate.objects.filter(date=start_date).delete()
                logger.info("Removed network hash rate info")
                BlocksFoundByPoolPerWindow.objects.filter(window_start_date=start_date).delete()
                logger.info("Removed block window info")
                BitcoinDifficulty.objects.filter(date=start_date).delete()
                logger.info("Removed bitcoin difficulty info")

                MetricsCalculationRunner.calculate_metrics_for_date_range(
                    start_date=start_date,
                    end_date=end_date
                )
                self.message_user(request, "The calculations for the date range have been finished!")
                return redirect("..")
                
            form = DateRangeForm()
            payload = {"form": form}
            return render(
                request, "admin/date_range_emission_calc_form.html", payload
            )
        else:
            return redirect("../..")

    
    def clear_all_data(self, request):
        if request.user.is_authenticated:   
            PoolElectricityConsumptionAndCO2EEmissionHistory.objects.all().delete()
            logger.info("Removed all location emission history")
            CO2ElectricityHistoryPerServer.objects.all().delete()
            logger.info("Removed all pool emission history")
            HashRatePerPoolServer.objects.all().delete()
            logger.info("Removed hash rate per server info")
            AverageEfficiency.objects.all().delete()
            logger.info("Removed average efficiency info")
            NetworkHashRate.objects.all().delete()
            logger.info("Removed network hash rate info")
            BlocksFoundByPoolPerWindow.objects.all().delete()
            logger.info("Removed all window info")
            PoolLocation.objects.all().delete()
            logger.info("Removed pool location window info")
            Location.objects.all().delete()
            logger.info("Removed location  info")
            Pool.objects.all().delete()
            logger.info("Removed pool info")
            BitcoinDifficulty.objects.all().delete()
            logger.info("Removed bitcoin difficulty info")
            MiningGear.objects.all().delete()
            logger.info("Removed mining gear info")
            UUIDModel.objects.all().delete()
            logger.info("Removed UUID info")
            self.message_user(request, "All data has been cleared from the database. Time to start fresh!")
            return redirect("..")
        else:
            return redirect("../..")
    
    def get_ordering(self, request):
        return ['-date', Lower('location_of_servers__location_name')]


@admin.register(CO2ElectricityHistoryPerServer)
class PoolEmissionHistoryAdmin(admin.ModelAdmin):
    date_hierarchy = 'date'

    def get_ordering(self, request):
        return ['-date', Lower('server_info__blockchain_pool__pool_name')]


@admin.register(PoolLocation)
class PoolLocationAdmin(admin.ModelAdmin):
    change_list_template = "entities/change_list_import_excel.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.import_excel),
        ]
        return my_urls + urls

    
    def import_excel(self, request):
        if request.user.is_authenticated:
            if request.method == "POST":
                excel_file = request.FILES["excel_file"]
                pool_data = load_workbook(excel_file.file)
                # Create database objects from passed in data
                ExcelParser.parse_excel_for_pool_and_location_info(workbook=pool_data)
                self.message_user(request, "Your Excel file has been imported")
                return redirect("..")
            form = ExcelImportForm()
            payload = {"form": form}
            return render(
                request, "admin/excel_form.html", payload
            )
        else:
            return redirect("../../") 


    def get_ordering(self, request):
        return ['-valid_for_date', Lower('blockchain_pool__pool_name')]
        

@admin.register(NetworkHashRate)
class NetworkHashrateAdmin(admin.ModelAdmin):
    date_hierarchy = 'date'

    def get_ordering(self, request):
        return ['-date']


@admin.register(BitcoinDifficulty)
class BitcoinDifficultyAdmin(admin.ModelAdmin):
    date_hierarchy = 'date'

    def get_ordering(self, request):
        return ['-date']
    
@admin.register(AverageEfficiency)
class AverageEfficiencyAdmin(admin.ModelAdmin):
    date_hierarchy = 'date'

    def get_ordering(self, request):
        return ['-date']